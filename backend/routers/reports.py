"""
Report Router - HR Absence Reports
Generates comprehensive absence reports with Bradford Factor and Pattern Detection.
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract
from typing import List, Optional
from datetime import datetime, timedelta
from pydantic import BaseModel
import io

from database import get_db, Employee, LeaveRequest, Department, User
from security import get_current_user

router = APIRouter(prefix="/reports", tags=["Reports"])

# --- SCHEMAS ---

class AbsenceReportRequest(BaseModel):
    start_date: str  # YYYY-MM-DD
    end_date: str    # YYYY-MM-DD
    employee_ids: Optional[List[int]] = None
    department_id: Optional[int] = None
    leave_types: Optional[List[str]] = None  # vacation, sick, permit, etc.

class EmployeeAbsenceSummary(BaseModel):
    employee_id: int
    employee_name: str
    department_name: str
    total_days: int
    absence_count: int  # Number of separate absence episodes
    bradford_factor: int
    by_type: dict  # { "FERIE": 5, "MALATTIA": 3, ... }
    monday_friday_count: int  # Pattern detection

class DepartmentSummary(BaseModel):
    department_id: int
    department_name: str
    total_days: int
    employee_count: int
    avg_days_per_employee: float

class AbsenceReportResponse(BaseModel):
    period: dict  # { start, end, days }
    employees: List[EmployeeAbsenceSummary]
    departments: List[DepartmentSummary]
    totals: dict
    patterns: dict  # Monday/Friday analysis
    year_comparison: Optional[dict] = None


# --- HELPER FUNCTIONS ---

def calculate_bradford_factor(absence_count: int, total_days: int) -> int:
    """
    Bradford Factor = S² × D
    S = number of absence episodes
    D = total days absent
    
    The Bradford Factor is used by HR departments worldwide to measure the
    impact of employee absences. It weighs frequent short absences more heavily
    than occasional long ones because they cause more disruption.
    
    Example:
    - 1 absence of 10 days = 1² × 10 = 10
    - 10 absences of 1 day = 10² × 10 = 1000
    
    Guidelines:
    - 0-49: Normal
    - 50-124: Attention needed
    - 125-399: Action required
    - 400+: Critical
    """
    return (absence_count ** 2) * total_days


def is_monday_or_friday(date_obj) -> bool:
    """Check if date is Monday (0) or Friday (4)"""
    return date_obj.weekday() in [0, 4]


def get_leave_days(start_date: datetime, end_date: datetime) -> int:
    """Calculate number of days between dates (inclusive)"""
    return (end_date - start_date).days + 1


# --- ENDPOINTS ---

@router.post("/absences", response_model=AbsenceReportResponse, summary="Generate Absence Report")
async def generate_absence_report(
    request: AbsenceReportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generates a comprehensive absence report with:
    - Per-employee breakdown
    - Per-department summary
    - Bradford Factor calculation
    - Monday/Friday pattern detection
    - Year-over-year comparison
    """
    
    # Access Control: Only super_admin for now
    if current_user.role != 'super_admin':
        raise HTTPException(status_code=403, detail="Accesso non autorizzato ai report")
    
    # Parse dates
    try:
        s_date = datetime.strptime(request.start_date, "%Y-%m-%d")
        e_date = datetime.strptime(request.end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato data non valido")
    
    period_days = (e_date - s_date).days + 1
    
    # Build employee filter
    emp_query = db.query(Employee).filter(Employee.is_active == True)
    
    if request.employee_ids:
        emp_query = emp_query.filter(Employee.id.in_(request.employee_ids))
    if request.department_id:
        emp_query = emp_query.filter(Employee.department_id == request.department_id)
    
    employees = emp_query.all()
    employee_ids = [e.id for e in employees]
    
    if not employee_ids:
        return AbsenceReportResponse(
            period={"start": request.start_date, "end": request.end_date, "days": period_days},
            employees=[],
            departments=[],
            totals={"total_days": 0, "total_episodes": 0, "avg_bradford": 0},
            patterns={"monday_count": 0, "friday_count": 0, "suspicious_rate": 0}
        )
    
    # Get Leave Requests in period
    leaves_query = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id.in_(employee_ids),
        LeaveRequest.status == 'approved',
        LeaveRequest.start_date <= e_date,
        LeaveRequest.end_date >= s_date
    )
    
    if request.leave_types:
        leaves_query = leaves_query.filter(LeaveRequest.leave_type.in_(request.leave_types))
    
    leaves = leaves_query.all()
    
    # Process data per employee
    emp_data = {}  # employee_id -> { ... }
    
    for emp in employees:
        emp_data[emp.id] = {
            "employee_id": emp.id,
            "employee_name": f"{emp.last_name} {emp.first_name}",
            "department_name": emp.department.name if emp.department else "N/D",
            "department_id": emp.department_id,
            "total_days": 0,
            "absence_count": 0,
            "by_type": {},
            "monday_friday_count": 0,
            "leave_dates": []
        }
    
    for leave in leaves:
        emp_id = leave.employee_id
        if emp_id not in emp_data:
            continue
        
        # Calculate overlap with requested period
        actual_start = max(leave.start_date, s_date)
        actual_end = min(leave.end_date, e_date)
        days = (actual_end - actual_start).days + 1
        
        emp_data[emp_id]["total_days"] += days
        emp_data[emp_id]["absence_count"] += 1
        
        # By type
        leave_type = leave.leave_type.upper()
        if leave_type == 'VACATION': leave_type = 'FERIE'
        elif leave_type == 'SICK': leave_type = 'MALATTIA'
        elif leave_type == 'PERMIT': leave_type = 'PERMESSO'
        
        emp_data[emp_id]["by_type"][leave_type] = emp_data[emp_id]["by_type"].get(leave_type, 0) + days
        
        # Pattern detection: Check if start or end is Mon/Fri
        if is_monday_or_friday(actual_start) or is_monday_or_friday(actual_end):
            emp_data[emp_id]["monday_friday_count"] += 1
    
    # Build employee summaries
    employee_summaries = []
    total_days_all = 0
    total_episodes_all = 0
    total_bradford = 0
    total_mon_fri = 0
    
    for emp_id, data in emp_data.items():
        bradford = calculate_bradford_factor(data["absence_count"], data["total_days"])
        
        employee_summaries.append(EmployeeAbsenceSummary(
            employee_id=data["employee_id"],
            employee_name=data["employee_name"],
            department_name=data["department_name"],
            total_days=data["total_days"],
            absence_count=data["absence_count"],
            bradford_factor=bradford,
            by_type=data["by_type"],
            monday_friday_count=data["monday_friday_count"]
        ))
        
        total_days_all += data["total_days"]
        total_episodes_all += data["absence_count"]
        total_bradford += bradford
        total_mon_fri += data["monday_friday_count"]
    
    # Sort by Bradford Factor (highest first)
    employee_summaries.sort(key=lambda x: x.bradford_factor, reverse=True)
    
    # Department Summaries
    dept_data = {}
    for emp_id, data in emp_data.items():
        dept_id = data["department_id"]
        if dept_id not in dept_data:
            dept_data[dept_id] = {
                "department_id": dept_id,
                "department_name": data["department_name"],
                "total_days": 0,
                "employee_count": 0
            }
        dept_data[dept_id]["total_days"] += data["total_days"]
        dept_data[dept_id]["employee_count"] += 1
    
    department_summaries = []
    for dept_id, data in dept_data.items():
        avg = data["total_days"] / data["employee_count"] if data["employee_count"] > 0 else 0
        department_summaries.append(DepartmentSummary(
            department_id=data["department_id"] or 0,
            department_name=data["department_name"],
            total_days=data["total_days"],
            employee_count=data["employee_count"],
            avg_days_per_employee=round(avg, 1)
        ))
    
    department_summaries.sort(key=lambda x: x.total_days, reverse=True)
    
    # Year-over-year comparison
    year_comparison = None
    prev_year_start = s_date.replace(year=s_date.year - 1)
    prev_year_end = e_date.replace(year=e_date.year - 1)
    
    prev_leaves = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id.in_(employee_ids),
        LeaveRequest.status == 'approved',
        LeaveRequest.start_date <= prev_year_end,
        LeaveRequest.end_date >= prev_year_start
    ).all()
    
    prev_total_days = 0
    for leave in prev_leaves:
        actual_start = max(leave.start_date, prev_year_start)
        actual_end = min(leave.end_date, prev_year_end)
        prev_total_days += (actual_end - actual_start).days + 1
    
    if prev_total_days > 0:
        change_pct = round(((total_days_all - prev_total_days) / prev_total_days) * 100, 1)
        year_comparison = {
            "previous_period": f"{prev_year_start.strftime('%d/%m/%Y')} - {prev_year_end.strftime('%d/%m/%Y')}",
            "previous_days": prev_total_days,
            "current_days": total_days_all,
            "change_percent": change_pct,
            "trend": "up" if change_pct > 0 else "down" if change_pct < 0 else "stable"
        }
    
    # Pattern analysis
    suspicious_rate = round((total_mon_fri / total_episodes_all * 100), 1) if total_episodes_all > 0 else 0
    
    return AbsenceReportResponse(
        period={"start": request.start_date, "end": request.end_date, "days": period_days},
        employees=employee_summaries,
        departments=department_summaries,
        totals={
            "total_days": total_days_all,
            "total_episodes": total_episodes_all,
            "avg_bradford": round(total_bradford / len(employee_summaries), 0) if employee_summaries else 0,
            "employees_analyzed": len(employee_summaries)
        },
        patterns={
            "monday_friday_count": total_mon_fri,
            "total_episodes": total_episodes_all,
            "suspicious_rate": suspicious_rate
        },
        year_comparison=year_comparison
    )


@router.post("/absences/export/excel", summary="Export Absence Report to Excel")
async def export_absence_report_excel(
    request: AbsenceReportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export detailed absence report as Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Access Control
    if current_user.role != 'super_admin':
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Get report data (reuse logic)
    report_data = await generate_absence_report(request, db, current_user)
    
    # Get detailed leaves for Excel
    s_date = datetime.strptime(request.start_date, "%Y-%m-%d")
    e_date = datetime.strptime(request.end_date, "%Y-%m-%d")
    
    emp_query = db.query(Employee).filter(Employee.is_active == True)
    if request.employee_ids:
        emp_query = emp_query.filter(Employee.id.in_(request.employee_ids))
    if request.department_id:
        emp_query = emp_query.filter(Employee.department_id == request.department_id)
    employees = emp_query.all()
    employee_ids = [e.id for e in employees]
    emp_map = {e.id: e for e in employees}
    
    leaves = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id.in_(employee_ids),
        LeaveRequest.status == 'approved',
        LeaveRequest.start_date <= e_date,
        LeaveRequest.end_date >= s_date
    ).order_by(LeaveRequest.start_date).all()
    
    # Create Workbook
    wb = Workbook()
    
    # --- SHEET 1: Dettaglio Assenze ---
    ws1 = wb.active
    ws1.title = "Dettaglio Assenze"
    
    # Header styling
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")  # Slate 900
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers1 = ["Dipendente", "Reparto", "Tipo", "Data Inizio", "Data Fine", "Giorni", "Motivo"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    row = 2
    for leave in leaves:
        emp = emp_map.get(leave.employee_id)
        if not emp:
            continue
        
        leave_type = leave.leave_type.upper()
        if leave_type == 'VACATION': leave_type = 'FERIE'
        elif leave_type == 'SICK': leave_type = 'MALATTIA'
        elif leave_type == 'PERMIT': leave_type = 'PERMESSO'
        
        actual_start = max(leave.start_date, s_date)
        actual_end = min(leave.end_date, e_date)
        days = (actual_end - actual_start).days + 1
        
        ws1.cell(row=row, column=1, value=f"{emp.last_name} {emp.first_name}")
        ws1.cell(row=row, column=2, value=emp.department.name if emp.department else "N/D")
        ws1.cell(row=row, column=3, value=leave_type)
        ws1.cell(row=row, column=4, value=actual_start.strftime("%d/%m/%Y"))
        ws1.cell(row=row, column=5, value=actual_end.strftime("%d/%m/%Y"))
        ws1.cell(row=row, column=6, value=days)
        ws1.cell(row=row, column=7, value=leave.reason or "-")
        row += 1
    
    # Auto-width columns
    for col in range(1, 8):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    
    # --- SHEET 2: Riepilogo Dipendenti ---
    ws2 = wb.create_sheet("Riepilogo Dipendenti")
    
    headers2 = ["Dipendente", "Reparto", "Giorni Totali", "N. Episodi", "Bradford Factor", "Lun/Ven %"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    row = 2
    for emp in report_data.employees:
        ws2.cell(row=row, column=1, value=emp.employee_name)
        ws2.cell(row=row, column=2, value=emp.department_name)
        ws2.cell(row=row, column=3, value=emp.total_days)
        ws2.cell(row=row, column=4, value=emp.absence_count)
        ws2.cell(row=row, column=5, value=emp.bradford_factor)
        mon_fri_pct = round((emp.monday_friday_count / emp.absence_count * 100), 0) if emp.absence_count > 0 else 0
        ws2.cell(row=row, column=6, value=f"{mon_fri_pct}%")
        row += 1
    
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    
    # --- SHEET 3: Riepilogo Reparti ---
    ws3 = wb.create_sheet("Riepilogo Reparti")
    
    headers3 = ["Reparto", "Giorni Totali", "Dipendenti", "Media GG/Dip"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 2
    for dept in report_data.departments:
        ws3.cell(row=row, column=1, value=dept.department_name)
        ws3.cell(row=row, column=2, value=dept.total_days)
        ws3.cell(row=row, column=3, value=dept.employee_count)
        ws3.cell(row=row, column=4, value=dept.avg_days_per_employee)
        row += 1
    
    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 20
    
    # --- SHEET 4: Bradford Factor Info ---
    ws4 = wb.create_sheet("Info Bradford Factor")
    ws4.cell(row=1, column=1, value="COS'È IL BRADFORD FACTOR?").font = Font(bold=True, size=14)
    ws4.cell(row=3, column=1, value="Formula: Bradford = S² × D")
    ws4.cell(row=4, column=1, value="S = Numero di episodi di assenza")
    ws4.cell(row=5, column=1, value="D = Giorni totali di assenza")
    ws4.cell(row=7, column=1, value="PERCHÉ È UTILE?").font = Font(bold=True)
    ws4.cell(row=8, column=1, value="Le assenze brevi e frequenti causano più disagi di una lunga.")
    ws4.cell(row=9, column=1, value="Chi si assenta 10 volte per 1 giorno ha Bradford 1000.")
    ws4.cell(row=10, column=1, value="Chi si assenta 1 volta per 10 giorni ha Bradford 10.")
    ws4.cell(row=12, column=1, value="SOGLIE DI ATTENZIONE:").font = Font(bold=True)
    ws4.cell(row=13, column=1, value="0-49: Normale")
    ws4.cell(row=14, column=1, value="50-124: Da monitorare")
    ws4.cell(row=15, column=1, value="125-399: Richiede azione")
    ws4.cell(row=16, column=1, value="400+: Critico")
    ws4.column_dimensions['A'].width = 60
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Report_Assenze_{request.start_date}_{request.end_date}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
